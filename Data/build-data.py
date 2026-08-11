# -*- coding: utf-8 -*-
"""
Convierte data.xlsx -> data.js (filas a nivel de obra para el dashboard).
Ejecutar cada vez que se actualice el Excel:  python build-data.py
(o doble clic en actualizar-datos.bat)
"""
import json, os, re, unicodedata
from datetime import datetime
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(BASE, "data.xlsx")
OUT = os.path.join(BASE, "data.js")

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb.active
hdr = [c.value for c in ws[1]]
idx = {h: i for i, h in enumerate(hdr)}

def col_like(prefix):
    for h in hdr:
        if h and str(h).upper().startswith(prefix):
            return h
    raise KeyError(prefix)

COL_PROY = col_like("PROYECCION")
COL_DET = col_like("DETALLE DEL MOTIVO")

def cell(r, col):
    v = r[idx[col]]
    return v.strip() if isinstance(v, str) else v

def num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0

def noacc(s):
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

def sent(s):
    s = ' '.join(str(s).split())
    return s[:1].upper() + s[1:].lower() if s else s

MOTIVO = {
    'INCUMPLIMIENTO CONTRACTUAL': 'Incumplimiento contractual',
    'DEFICIENCIA DEL EXPEDIENTE TECNICO': 'Deficiencia del expediente técnico',
    'DEFICIENCIAS ADMINISTRATIVAS': 'Deficiencias administrativas',
    'CONFLICTOS SOCIALES': 'Conflictos sociales',
    'NULIDAD DEL CONTRATO': 'Nulidad del contrato',
    'DEFICIENCIAS CONSTRUCTIVAS': 'Deficiencias constructivas',
    'POR APLICACIÓN DE LA CLÁUSULA ANTICORRUPCIÓN': 'Cláusula anticorrupción',
    'FACTORES CLIMATOLOGICOS': 'Factores climatológicos',
}
HITO = {
    'DIAGNOSTICO': 'Diagnóstico',
    'INFORME DE ESTADO SITUACIONAL': 'Informe de estado situacional',
    'E.T. SALDO DE OBRA': 'E.T. saldo de obra',
    'PROCESO DE SELECCIÓN DEL SALDO DE OBRA': 'Proceso de selección',
    'REACTIVADO': 'Reactivado',
}
NIVEL = {'LOCAL': 'Local', 'NACIONAL': 'Nacional', 'REGIONAL': 'Regional', 'EPS': 'EPS'}
TRIM = {'I TRIMESTRE': 'I Trim 26', 'II TRIMESTRE': 'II Trim 26',
        'III TRIMESTRE': 'III Trim 26', 'IV TRIMESTRE': 'IV Trim 26'}

SSP = {'CONCLUIDO': 'Concluido', 'EN EJECUCIÓN': 'En ejecución', 'PARALIZADA': 'Paralizada'}

rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    est = str(cell(r, 'ESTADO DE LA OBRA') or '').strip().upper()
    if est not in ('PARALIZADO', 'REACTIVADO'):
        continue
    p = str(cell(r, COL_PROY) or '').strip().upper()
    py = TRIM.get(p, p if p.isdigit() else None)
    f = cell(r, 'FECHA DE ULTIMA PARALIZACION')
    if isinstance(f, datetime):
        ap = f.year
    else:
        s = str(f or '')[:4]
        ap = int(s) if s.isdigit() else None
    ni = str(cell(r, 'NIVEL DE GOBIERNO') or '').strip().upper()
    mo_raw = str(cell(r, 'MOTIVO DE PARALIZACION') or '').strip().upper()
    hi_raw = str(cell(r, 'HITO ACTUAL') or '').strip().upper()
    mo2_raw = str(cell(r, 'MOTIVO DE SEGUNDA PARALIZACION') or '').strip()
    mo2 = None
    if mo2_raw and mo2_raw != '-':
        mo2 = re.sub(r'\bet\b', 'E.T.', sent(mo2_raw))
    ssp_raw = str(cell(r, 'ESTADO SSP') or '').strip().upper()
    sub_raw = str(cell(r, 'SUB ESTADO SSP') or '').strip().upper()
    frd = cell(r, 'FECHA REAL DE REACTIVACIÓN')
    fr = frd.strftime('%Y-%m') if isinstance(frd, datetime) else None
    fpd = cell(r, 'FECHA DE ULTIMA PARALIZACION')
    mp = None
    if isinstance(frd, datetime) and isinstance(fpd, datetime):
        mp = (frd.year - fpd.year) * 12 + (frd.month - fpd.month)
    rows.append({
        'e': 'P' if est == 'PARALIZADO' else 'R',
        'c': cell(r, 'CUI'),
        'n': sent(cell(r, 'NOMBRE DEL PROYECTO') or '')[:120],
        'd': noacc(str(cell(r, 'DEPARTAMENTO') or '').strip().upper()),
        'ni': NIVEL.get(ni, ni.title()),
        's': str(cell(r, 'SECTOR') or '').strip().title(),
        'mf': str(cell(r, 'MODALIDAD DE FINANCIAMIENTO') or '').strip().upper(),
        'pg': str(cell(r, 'PROG.') or '').strip(),
        'av': str(cell(r, 'RANGO % AVANCE FISICO') or '').strip().replace('-', '–'),
        'mo': MOTIVO.get(mo_raw, sent(mo_raw)),
        'de': sent(cell(r, COL_DET) or ''),
        'ap': ap,
        'hi': HITO.get(hi_raw, sent(hi_raw)),
        'py': py,
        'm26': 1 if str(cell(r, 'META 2026') or '').strip() == 'META REPORTE' else 0,
        'ssp': SSP.get(ssp_raw, sent(ssp_raw)),
        'm2': 1 if (ssp_raw == 'PARALIZADA' or 'SUSPEN' in sub_raw) else 0,
        'mo2': mo2,
        'fr': fr,
        'mp': mp,
        'po': int(num(cell(r, 'POBLACION BENEFICIARIA'))),
        'ci': round(num(cell(r, 'COSTO DE INVERSIÓN')), 2),
        'dv': round(num(cell(r, 'DEVENGADO ACUMULADO')), 2),
        'pim': round(num(cell(r, 'PIM 2025')), 2),
        'd25': round(num(cell(r, 'DEVENGADO 2025')), 2),
    })

with open(OUT, 'w', encoding='utf-8') as fh:
    fh.write('/* Generado desde data.xlsx por build-data.py — NO EDITAR A MANO */\n')
    fh.write('const ROWS = ' + json.dumps(rows, ensure_ascii=False, separators=(',', ':')) + ';\n')

print('OK:', len(rows), 'obras ->', OUT)
p = sum(1 for x in rows if x['e'] == 'P')
print('Paralizadas:', p, '| Reactivadas:', len(rows) - p)
